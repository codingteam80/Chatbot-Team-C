import csv
import html
import json
import re
import time
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components

try:
    from docx import Document
    from docx.enum.text import WD_ALIGN_PARAGRAPH
    from docx.shared import Inches as DocxInches
    from docx.shared import Pt as DocxPt
except Exception:
    Document = None
    WD_ALIGN_PARAGRAPH = None
    DocxInches = None
    DocxPt = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except Exception:
    Workbook = None
    Alignment = None
    Border = None
    Font = None
    PatternFill = None
    Side = None

try:
    from pptx import Presentation
    from pptx.util import Inches as PptxInches
    from pptx.util import Pt as PptxPt
except Exception:
    Presentation = None
    PptxInches = None
    PptxPt = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
except Exception:
    colors = None
    letter = None
    ParagraphStyle = None
    getSampleStyleSheet = None
    Paragraph = None
    SimpleDocTemplate = None
    Spacer = None

from ui.source_ui import (
    clean_preview_text,
    deduplicate_sources,
    make_static_file_link,
    normalize_source_item,
)

SIDEBAR_SOURCES_KEY = "sidebar_sources"
SOURCE_PREVIEW_LIMIT = 360
EXPORT_TIMESTAMP_FORMAT = "%Y%m%d_%H%M%S"
PENDING_EXPORT_REQUEST_KEY = "pending_export_request"
READY_EXPORT_FILE_KEY = "ready_export_file"
EXPORT_FRONTEND_FLUSH_SECONDS = 0.12
ACTION_DEBUG_FILE = Path(__file__).resolve().parents[1] / "reports" / "ui_rag_debug.txt"


def append_action_debug(title, lines=None):
    # Confirms whether action_ui.py receives the button click.
    try:
        ACTION_DEBUG_FILE.parent.mkdir(parents=True, exist_ok=True)
        log_lines = [
            "",
            "=" * 80,
            str(title or "ACTION DEBUG"),
            "=" * 80,
            f"Time: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        ]

        for line in lines or []:
            log_lines.append(str(line))

        with ACTION_DEBUG_FILE.open("a", encoding="utf-8") as file:
            file.write("\n".join(log_lines))
            file.write("\n")

        print(f"ACTION DEBUG: {title} -> {ACTION_DEBUG_FILE.resolve()}", flush=True)
    except Exception as error:
        print(f"ACTION DEBUG WRITE FAILED: {error}", flush=True)


print(f"ACTIVE ACTION_UI FILE: {Path(__file__).resolve()}", flush=True)
print(f"ACTION DEBUG FILE TARGET: {ACTION_DEBUG_FILE.resolve()}", flush=True)

EXPORT_TYPES = {
    "PDF": ("pdf", "application/pdf"),
    "DOCX": ("docx", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
    "TXT": ("txt", "text/plain"),
    "Markdown": ("md", "text/markdown"),
    "CSV": ("csv", "text/csv"),
    "XLSX": ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "XLS": ("xls", "application/vnd.ms-excel"),
    "PPTX": ("pptx", "application/vnd.openxmlformats-officedocument.presentationml.presentation"),
    "PPT": ("ppt", "application/vnd.ms-powerpoint"),
}


def safe_html(value):
    # Escape text for custom HTML blocks.
    return html.escape(str(value or ""), quote=True)


def copy_button(text, key):
    # Browser clipboard needs a tiny component iframe; main CSS cannot style inside iframes.
    safe_text = json.dumps(text or "")
    safe_key = safe_html(key)

    components.html(
        f"""
<!DOCTYPE html>
<html>
<head>
<style>
html,body{{margin:0;padding:0;width:34px;height:34px;overflow:hidden;background:transparent;}}
button{{width:34px;height:34px;border:0;border-radius:10px;background:transparent;color:rgba(255,255,255,.82);font-size:15px;cursor:pointer;}}
button:hover{{background:rgba(255,255,255,.08);color:#fff;}}
</style>
</head>
<body>
<button id="copy-btn-{safe_key}" title="Copy" onclick="copyText()">⧉</button>
<script>
function fallbackCopy(text){{
  const el=document.createElement('textarea');
  el.value=text; el.style.position='fixed'; el.style.opacity='0';
  document.body.appendChild(el); el.focus(); el.select();
  try{{document.execCommand('copy');}}catch(err){{}}
  document.body.removeChild(el);
}}
function showCopied(){{
  const btn=document.getElementById('copy-btn-{safe_key}');
  btn.innerText='✓'; setTimeout(function(){{btn.innerText='⧉';}},900);
}}
function copyText(){{
  const text={safe_text};
  if(navigator.clipboard && window.isSecureContext){{
    navigator.clipboard.writeText(text).then(showCopied).catch(function(){{fallbackCopy(text);showCopied();}});
  }} else {{ fallbackCopy(text); showCopied(); }}
}}
</script>
</body>
</html>
""",
        height=34,
        width=34,
        scrolling=False,
    )


def get_export_timestamp():
    # One timestamp per click/rerun for readable filenames and export metadata.
    return datetime.now().strftime(EXPORT_TIMESTAMP_FORMAT)


def title_case_words(text):
    # Make a readable title without depending on the LLM.
    small_words = {"a", "an", "and", "as", "at", "by", "for", "from", "in", "of", "on", "or", "the", "to", "with"}
    words = str(text or "").split()
    result = []

    for index, word in enumerate(words):
        if word.isupper() and len(word) <= 6:
            result.append(word)
            continue

        lower_word = word.lower()

        if index > 0 and lower_word in small_words:
            result.append(lower_word)
        else:
            result.append(lower_word[:1].upper() + lower_word[1:])

    return " ".join(result).strip()


def make_response_title(question=None):
    # Create a short topic title related to the question, but not the exact question.
    text = str(question or "").strip()

    if not text:
        return "InknowVa Response"

    text = re.sub(r"[?!.]+$", "", text)
    text = re.sub(r"\s+", " ", text).strip()

    replacements = [
        (r"^(please\s+)?(can|could|would)\s+you\s+", ""),
        (r"^(please\s+)?tell\s+me\s+about\s+", ""),
        (r"^(please\s+)?explain\s+", ""),
        (r"^what\s+is\s+", ""),
        (r"^what\s+are\s+", ""),
        (r"^who\s+is\s+", ""),
        (r"^who\s+are\s+", ""),
        (r"^when\s+did\s+", ""),
        (r"^when\s+is\s+", ""),
        (r"^where\s+is\s+", ""),
        (r"^where\s+are\s+", ""),
        (r"^why\s+is\s+", ""),
        (r"^why\s+are\s+", ""),
        (r"^how\s+does\s+", ""),
        (r"^how\s+do\s+", ""),
        (r"^how\s+did\s+", ""),
        (r"^ano\s+ang\s+", ""),
        (r"^sino\s+si\s+", ""),
        (r"^sino\s+ang\s+", ""),
        (r"^kailan\s+", ""),
        (r"^bakit\s+", ""),
        (r"^paano\s+", ""),
    ]

    topic = text
    for pattern, replacement in replacements:
        topic = re.sub(pattern, replacement, topic, flags=re.IGNORECASE).strip()

    topic = re.sub(r"\b(about|regarding|related to)\b", "", topic, flags=re.IGNORECASE).strip()
    topic = re.sub(r"\s+", " ", topic).strip(" -:;,")

    if not topic or topic.lower() == text.lower():
        topic = text

    # Avoid using the exact question wording as a document title.
    if topic.lower() == text.lower():
        title = f"{title_case_words(topic)} Summary"
    else:
        title = title_case_words(topic)

    if len(title) > 70:
        title = title[:70].rsplit(" ", 1)[0].strip()

    return title or "InknowVa Response"


def make_filename_slug(title):
    # Safe, short filename part based on the generated topic title.
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", str(title or "inknowva_response").lower()).strip("_")
    slug = re.sub(r"_+", "_", slug)
    return slug[:60].strip("_") or "inknowva_response"


def make_export_filename(export_type, title=None):
    # Dynamic filename based on the topic title, not the exact question.
    extension, _mime = EXPORT_TYPES[export_type]
    slug = make_filename_slug(title or "InknowVa Response")
    return f"{slug}_{get_export_timestamp()}.{extension}"


def get_question_for_message(message_index=None):
    # Pull the original question from the assistant message when available.
    if message_index is None:
        return ""

    messages = st.session_state.get("messages", [])

    try:
        message = messages[int(message_index)]
    except Exception:
        return ""

    if isinstance(message, dict):
        question = str(message.get("question") or "").strip()

        if question:
            return question

    # Fallback: use nearest previous user message.
    try:
        start_index = int(message_index) - 1
    except Exception:
        return ""

    for index in range(start_index, -1, -1):
        previous_message = messages[index]

        if isinstance(previous_message, dict) and previous_message.get("role") == "user":
            return str(previous_message.get("content") or "").strip()

    return ""


def normalize_answer_text(answer):
    # Normalize answer text while preserving paragraphs and list lines.
    text = str(answer or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text


def is_bullet_line(line):
    # Detect bullets from Markdown/plain text.
    return bool(re.match(r"^\s*([-*•])\s+", str(line or "")))


def is_numbered_line(line):
    # Detect numbered list items such as "1. item" or "1) item".
    return bool(re.match(r"^\s*\d+[\.\)]\s+", str(line or "")))


def strip_list_marker(line):
    # Remove bullet/number markers before applying document list styles.
    line = str(line or "").strip()
    line = re.sub(r"^\s*[-*•]\s+", "", line)
    line = re.sub(r"^\s*\d+[\.\)]\s+", "", line)
    return line.strip()


def split_answer_lines(answer):
    # Return meaningful lines while keeping blank lines as paragraph breaks.
    text = normalize_answer_text(answer)

    if not text:
        return []

    return text.split("\n")


def get_source_entries(sources):
    # Normalize and deduplicate sources for all export formats.
    entries = []

    for source in deduplicate_sources(sources or [], max_sources=None):
        item = normalize_source_item(source) or {}

        if not item:
            continue

        name = str(item.get("source") or "Unknown source").strip()
        page = str(item.get("page") or "N/A").strip()
        preview = clean_preview_text(item.get("preview"), limit=220)
        entries.append({
            "source": name,
            "page": page,
            "preview": preview,
        })

    return entries


def format_source_line(source, index=None):
    # Source line for export files.
    source = normalize_source_item(source) or {}
    prefix = f"{index}. " if index is not None else ""
    return f"{prefix}{source.get('source', 'Unknown source')} | Page: {source.get('page', 'N/A')}"


def build_export_text(answer, sources=None, title=None):
    # Shared plain-text body for text-style exports.
    # Exported files include only the topic title and answer; sources stay in the app UI.
    lines = [
        str(title or "InknowVa Response"),
        "",
        normalize_answer_text(answer) or "No answer.",
    ]

    return "\n".join(lines)


# Public names exported by this compatibility/refactor module.
__all__ = [
    'csv',
    'html',
    'json',
    're',
    'time',
    'datetime',
    'BytesIO',
    'StringIO',
    'Path',
    'st',
    'components',
    'Document',
    'WD_ALIGN_PARAGRAPH',
    'DocxInches',
    'DocxPt',
    'Workbook',
    'Alignment',
    'Border',
    'Font',
    'PatternFill',
    'Side',
    'Presentation',
    'PptxInches',
    'PptxPt',
    'colors',
    'letter',
    'ParagraphStyle',
    'getSampleStyleSheet',
    'Paragraph',
    'SimpleDocTemplate',
    'Spacer',
    'clean_preview_text',
    'deduplicate_sources',
    'make_static_file_link',
    'normalize_source_item',
    'SIDEBAR_SOURCES_KEY',
    'SOURCE_PREVIEW_LIMIT',
    'EXPORT_TIMESTAMP_FORMAT',
    'PENDING_EXPORT_REQUEST_KEY',
    'READY_EXPORT_FILE_KEY',
    'EXPORT_FRONTEND_FLUSH_SECONDS',
    'ACTION_DEBUG_FILE',
    'append_action_debug',
    'EXPORT_TYPES',
    'safe_html',
    'copy_button',
    'get_export_timestamp',
    'title_case_words',
    'make_response_title',
    'make_filename_slug',
    'make_export_filename',
    'get_question_for_message',
    'normalize_answer_text',
    'is_bullet_line',
    'is_numbered_line',
    'strip_list_marker',
    'split_answer_lines',
    'get_source_entries',
    'format_source_line',
    'build_export_text',
]
